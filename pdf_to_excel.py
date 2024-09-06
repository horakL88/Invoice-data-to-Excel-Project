import pdfquery
from openpyxl import Workbook

# Excel worksheetet aktiválom
wb = Workbook()
ws = wb.active

#reading pdf
pdf = pdfquery.PDFQuery(r'invoice.pdf')
pdf.load()

pages = pdf.doc.catalog['Pages'].resolve()['Count']
print(f"Total pages: {pages}")

#converting pdf to xml
pdf.tree.write(r'invoice.xml', pretty_print = True)
# Az alábbi listával csekkoljuk, hogy EU-s e egy értékesítés vagy sem
EU = ['AT', 'BE', 'BG', 'CY', 'CZ', 'DE', 'DK', 'DK', 'EE', 'EL', 'ES', 'FI', 'FR', 'HR', 'IE', 'IT', 'LT', 'LU', 'LV', 'MT', 'NL', 'PL', 'PT', 'RO', 'SE', 'SI', 'SK', 'XI']

# Fejlécek generálása
ws['A1'] = "Customer name" 
ws['B1'] = "Customer VAT status"
ws['C1'] = "Customer HU VAT Number" 
ws['D1'] = "Tax number of VAT group member" 
ws['E1'] = "CustomerCommunityVATNumber" 
ws['F1'] = "Customer Third country VAT Number (if any)" 
ws['G1'] = "Customer country" 
ws['H1'] = "Customer postal code" 
ws['I1'] = "Customer city/town" 
ws['J1'] = "Customer public place name" 
ws['K1'] = "Customer public place type" 
ws['L1'] = "Customer street no." 
ws['M1'] = "Customer Building" 
ws['N1'] = "Customer Stairway" 
ws['O1'] = "Customer Floor" 
ws['P1'] = "Customer Door" 
ws['Q1'] = "Fiscal representative's name (if any)" 
ws['R1'] = "Fiscal representative's tax ID (if any)" 
ws['S1'] = "Fiscal representative's country (if any)" 
ws['T1'] = "Fiscal representative's postal code (if any)" 
ws['U1'] = "Fiscal representative's city/town (if any)" 
ws['V1'] = "Fiscal representative's public place name (if any)" 
ws['W1'] = "Fiscal representative's public place type (if any)" 
ws['X1'] = "Fiscal representative's street no. (if any)" 
ws['Y1'] = "Fiscal representative's building (if any)" 
ws['Z1'] = "Fiscal representative's stairway (if any)" 
ws['AA1'] = "Fiscal representative's floor (if any)" 
ws['AB1'] = "Fiscal representative's door (if any)" 
ws['AC1'] = "Invoice number" 
ws['AD1'] = "Invoice type" 
ws['AE1'] = "Invoice issue date" 
ws['AF1'] = "Delivery date" 
ws['AG1'] = "Payment method" 
ws['AH1'] = "Currency" 
ws['AI1'] = "Exchange rate" 
ws['AJ1'] = "Invoice Appearance" 
ws['AK1'] = "Description" 
ws['AL1'] = "Quantity" 
ws['AM1'] = "Unit of measure" 
ws['AN1'] = "Unit of measure Own Description" 
ws['AO1'] = "Unit of measure definable?" 
ws['AP1'] = "Unit price" 
ws['AQ1'] = "Advance payment" 
ws['AR1'] = "Discount Description" 
ws['AS1'] = "Discount Amount" 
ws['AT1'] = "Discount %" 
ws['AU1'] = "Line Net amount" 
ws['AV1'] = "Tax rate" 
ws['AW1'] = "VAT percentage / Case" 
ws['AX1'] = "Reason" 
ws['AY1'] = "Line VAT amount" 
ws['AZ1'] = "Line Gross amount" 
ws['BA1'] = "Total Net amount" 
ws['BB1'] = "Total VAT amount" 
ws['BC1'] = "Total VAT amount in HUF" 
ws['BD1'] = "Total Gross amount"
ws['BE1'] = "Ship from?"

row = 2

# First lines description bbox details:

# A function that takes the data that is going to be the same for all lines inside of an invoice
def transfer(row):
    # Customer name
    customer_name = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'A{row}'] = customer_name
    # Customer VAT no
    customer_vat_no = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    if customer_vat_no[:2] in EU:
        ws[f'E{row}'] = customer_vat_no
        ws[f'B{row}'] = "Other (foreign taxpayer - Community, third country - domestic non-VAT subject, non-natural person,  and foreign non-VAT subject, non-natural person)"
    elif customer_vat_no[:2] == 'HU':
        ws[f'C{row}'] = customer_vat_no
        ws[f'B{row}'] = "Domestic taxpayer"
    # ws[f'B{row}'] = "Domestic or foreign natural person"
    # VAT group number
    VAT_group_no = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'D{row}'] = VAT_group_no
    # Customer third country VAT number
    customer_third_vat_no = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'F{row}'] = customer_third_vat_no
    
    # Customer adress data
    customer_country = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'G{row}'] = customer_country

    customer_postal_code = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'H{row}'] = customer_postal_code

    customer_city = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'I{row}'] = customer_city
    
    customer_public_place_name = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'J{row}'] = customer_public_place_name 
    
    customer_public_place_type = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'K{row}'] = customer_public_place_type  
    
    customer_street_no = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'L{row}'] = customer_street_no  
    
    customer_building = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'M{row}'] = customer_building 
    
    customer_stairway = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'N{row}'] = customer_stairway  
    
    customer_floor = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'O{row}'] = customer_floor  

    customer_room = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'P{row}'] = customer_room 

    ## Financial representative should be a constant if there is any
    ws[f'Q{row}'] = "Name"  
    ws[f'R{row}'] = "VAT number"
    ws[f'S{row}'] = "Country"
    ws[f'T{row}'] = "Postal code"  
    ws[f'U{row}'] = "City"   
    ws[f'V{row}'] = "Public Place name"      
    ws[f'W{row}'] = "Public place type"
    ws[f'X{row}'] = "Street no"
    ws[f'Y{row}'] = "Building" 
    ws[f'Z{row}'] = "Stairway"
    ws[f'AA{row}'] = "Floor"
    ws[f'AB{row}'] = "Door"

    # Invoice number 
    invoice_no = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'AC{row}'] = invoice_no
    # Invoice type
    ws[f'AD{row}'] = "NORMAL"
    ws[f'AD{row}'] = "SIMPLIFIED"
    ws[f'AD{row}'] = "AGGREGATE"
    # Invoice issue date
    issue_date = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'AE{row}'] = issue_date
    # Delivery Date
    delivery_date = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'AF{row}'] = delivery_date
    # Currency
    currency = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'AH{row}'] = currency
    # Exchange rate
    exchange_rate = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'AI{row}'] = exchange_rate
    # Invoice Appearance
    ws[f'AJ{row}'] = "Paper invoice"
    ws[f'AJ{row}'] = "Electronically created, non-EDI invoice"
    ws[f'AJ{row}'] = "EDI invoice"
    ws[f'AJ{row}'] = "The software cannot be identify the form of appearance of the invoice or it is unknownk at the time of issue"   
    # Total net amount
    total_net_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'BA{row}'] = total_net_amount
    # Total vat amount
    total_vat_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'BB{row}'] = total_vat_amount
    # Total vat amount in HUF
    total_net_amount_huf = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'BC{row}'] = total_net_amount_huf
    # Total gross amount
    total_gross_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
    ws[f'BD{row}'] = total_gross_amount
    total_gross_amount_check = total_net_amount + total_vat_amount
    if total_gross_amount_check != total_gross_amount:
        print (f"In row {row} the total gross amount on invoice may be wrong. Invoice: ", total_gross_amount_check, "Net+VAT = ", total_gross_amount)
    else: pass

for page in range(pages):
    pdf.load(page)
    # Description for first line
    description_x1 = 1
    description_y1 = 1
    description_x2 = 1
    description_y2 = 1
    description = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{description_x1}, {description_y1}, {description_x2}, {description_y2}")').text()
    # Quantity first line
    quantity_x1 = 1
    quantity_y1 = 1
    quantity_x2 = 1
    quantity_y2 = 1

    # Unit of measure first line
    unitofmeasure_x1 = 1
    unitofmeasure_y1 = 1
    unitofmeasure_x2 = 1
    unitofmeasure_y2 = 1

    # Unit of measure own description first line
    uom_own_description_x1 = 1
    uom_own_description_y1 = 1
    uom_own_description_x2 = 1
    uom_own_description_y2 = 1

    # Unit price first line
    unit_price_x1 = 1
    unit_price_y1 = 1
    unit_price_x2 = 1
    unit_price_y2 = 1

    # Discount Description first line
    discount_description_x1 = 1
    discount_description_y1 = 1
    discount_description_x2 = 1
    discount_description_y2 = 1

    # Discount Amount first line
    discount_amount_x1 = 1
    discount_amount_y1 = 1
    discount_amount_x2 = 1
    discount_amount_y2 = 1

    # Discount % first line
    discount_percent_x1 = 1
    discount_percent_y1 = 1
    discount_percent_x2 = 1
    discount_percent_y2 = 1

    # Line net amount first line
    line_net_amount_x1 = 1
    line_net_amount_y1 = 1
    line_net_amount_x2 = 1
    line_net_amount_y2 = 1

    # Tax rate first line
    tax_rate_x1 = 1
    tax_rate_y1 = 1
    tax_rate_x2 = 1
    tax_rate_y2 = 1

    # Line vat amount first line
    line_vat_amount_x1 = 1
    line_vat_amount_y1 = 1
    line_vat_amount_x2 = 1
    line_vat_amount_y2 = 1


    while description != None:
        # We can put extra checks here if needed
        transfer(row)
        # Description
        description = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{description_x1}, {description_y1}, {description_x2}, {description_y2}")').text()
        ws[f'AK{row}'] = description
        # We set for the next line based on position
        description_x1 = description_x1-1
        description_y1 = description_y1-1
        description_x2 = description_x2-1
        description_y2 = description_y2-1
        
        # Quantity
        quantity = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{quantity_x1}, {quantity_y1}, {quantity_x2}, {quantity_y2}")').text()
        ws[f'AL{row}'] = quantity
        # We set for the next line based on position
        quantity_x1 = quantity_x1-1
        quantity_y1 = quantity_y1-1
        quantity_x2 = quantity_x2-1
        quantity_y2 = quantity_y2-1
        
        # Unit of measure
        unit_of_measure = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{unitofmeasure_x1}, {unitofmeasure_y1}, {unitofmeasure_x2}, {unitofmeasure_y2}")').text()
        ws[f'AM{row}'] = unit_of_measure
        # We set for the next line based on position
        unitofmeasure_x1 = unitofmeasure_x1-1
        unitofmeasure_y1 = unitofmeasure_y1-1
        unitofmeasure_x2 = unitofmeasure_x2-1
        unitofmeasure_y2 = unitofmeasure_y1-1
        
        # Unit of measure own description
        unit_of_measure_own_description = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{uom_own_description_x1}, {uom_own_description_y1}, {uom_own_description_x2}, {uom_own_description_y2}")').text()
        ws[f'AN{row}'] = unit_of_measure_own_description
        # We set for the next line based on position
        uom_own_description_x1 = uom_own_description_x1-1
        uom_own_description_y1 = uom_own_description_y1-1
        uom_own_description_x2 = uom_own_description_x2-1
        uom_own_description_y2 = uom_own_description_y2-1
        
        # Unit of measure definable?
        if unit_of_measure and unit_of_measure_own_description is None:
            ws[f'AO{row}'] = "No"
        else:
            ws[f'AO{row}'] = "Yes"

        # Unit price
        unit_price = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{unit_price_x1}, {unit_price_y1}, {unit_price_x2}, {unit_price_y2}")').text()
        ws[f'AP{row}'] = unit_price
        # We set for the next line based on position
        unit_price_x1 = unit_price_x1-1
        unit_price_y1 = unit_price_y1-1
        unit_price_x2 = unit_price_x2-1
        unit_price_y2 = unit_price_y2-1

        # Advance payment
        ws[f'AQ{row}'] = "Yes"
        ws[f'AQ{row}'] = "No"

        # Discount Description
        discount_description = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{discount_description_x1}, {discount_description_y1}, {discount_description_x2}, {discount_description_y2}")').text()
        ws[f'AR{row}'] = discount_description
        # We set for the next line based on position
        discount_description_x1 = discount_description_x1-1
        discount_description_y1 = discount_description_y1-1
        discount_description_x2 = discount_description_x2-1
        discount_description_y2 = discount_description_y2-1

        # Discount amount
        discount_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{discount_amount_x1}, {discount_amount_y1}, {discount_amount_x2}, {discount_amount_y2}")').text()
        ws[f'AS{row}'] = discount_amount
        # We set for the next line based on position
        discount_amount_x1 = discount_amount_x1-1
        discount_amount_y1 = discount_amount_y1-1
        discount_amount_x2 = discount_amount_x2-1
        discount_amount_y2 = discount_amount_y2-1

        # Discount %
        discount_percent = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{discount_percent_x1}, {discount_percent_y1}, {discount_percent_x2}, {discount_percent_y2}")').text()
        ws[f'AT{row}'] = discount_percent
        # We set for the next line based on position
        discount_percent_x1 = discount_amount_x1-1
        discount_amount_y1 = discount_amount_y1-1
        discount_amount_x2 = discount_amount_x2-1
        discount_amount_y2 = discount_amount_y2-1

        # Line Net amount
        line_net_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{line_net_amount_x1}, {line_net_amount_y1}, {line_net_amount_x2}, {line_net_amount_y2}")').text()
        ws[f'AU{row}'] = line_net_amount
        # We set for the next line based on position
        line_net_amount_x1 = line_net_amount_x1-1
        line_net_amount_y1 = line_net_amount_y1-1
        line_net_amount_x2 = line_net_amount_x2-1
        line_net_amount_y2 = line_net_amount_y2-1

        # Tax rate
        tax_rate = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{tax_rate_x1}, {tax_rate_y1}, {tax_rate_x2}, {tax_rate_y2}")').text()
        ws[f'AV{row}'] = tax_rate
        # We set for the next line based on position
        tax_rate_x1 = tax_rate_x1-1
        tax_rate_y1 = tax_rate_y1-1
        tax_rate_x2 = tax_rate_x2-1
        tax_rate_y2 = tax_rate_y2-1

        # VAT percantage / Case
        ws[f'AW{row}'] = "Small entrepreneur exemption"
        ws[f'AW{row}'] = "VAT exemption based on nature of the supply"
        ws[f'AW{row}'] = "VAT exempt IC-supply (excl. new means of transport)"
        ws[f'AW{row}'] = "VAT exempt IC-supply of new means of transport"
        ws[f'AW{row}'] = "Export of goods to non-EU countries"
        ws[f'AW{row}'] = "Other exempt international transaction"
        ws[f'AW{row}'] = "-"

        # Reason
        ws[f'AX{row}'] = "Reason for VAT exemption if needed"

        # Line VAT amount
        line_vat_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("{line_vat_amount_x1}, {line_vat_amount_y1}, {line_vat_amount_x2}, {line_vat_amount_y2}")').text()
        ws[f'AY{row}'] = line_vat_amount
        # We set for the next line based on position
        line_vat_amount_x1 = line_vat_amount_x1-1
        line_vat_amount_y1 = line_vat_amount_y1-1
        line_vat_amount_x2 = line_vat_amount_x2-1
        line_vat_amount_y2 = line_vat_amount_y2-1

        # Line gross amount
        line_gross_amount = pdf.pq(f'LTTextLineHorizontal:overlaps_bbox("")').text()
        ws[f'AZ{row}'] = line_gross_amount
        line_gross_amount_check = line_net_amount + line_vat_amount
        if line_gross_amount_check != line_gross_amount:
            print (f"In row {row} the Line gross amount on invoice may be wrong. Invoice: ", line_gross_amount_check, "Net+VAT = ", line_gross_amount)
        else: pass
        
        row += 1


wb.save("invoice.xlsx")